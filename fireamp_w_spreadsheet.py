''' This script is utilized to query the fireAMP database

 This is a python 3 version of the script
 it will search against hash, file names, IP whatever the API takes
 
 Refer to the FireAMP API docs on how to generate the Base64 Encoded API key


 Tom Yarrish
 Version 0.6

 CHANGELOG

 7/1/2016

 1) Program will now determine the ending row, you don't need to enter that
 2) Uses new Client API Key
 3) If you get a hit that has more than 500 hosts, the program just prints
 out a message that there are too many hosts to list and moves on.
 4) When you get a hit, it now prints the list of hostnames to a text file
 per hit in the working directory
 5) Program prints out the remaining number of requests you can make to the
 FireAMP server
 6) Switched to base64 API request
 7) Rewrote functions since we don't need the API info with base64

 TODO

 1) Update script to v1 of API
 3) Query against VirusTotal?

'''
import requests
import argparse
import collections
from openpyxl import load_workbook

fireamp_base64_api = 'ENTER BASE64 API KEY HERE'

FireampResult = collections.namedtuple(
    'FireampResult',
    'active, connector_guid, hostname, links'
)


# This function processes the Excel file.
# You have to specify the worksheet tab to read,
# the column to process, and the starting and ending row.
def setup_excel_file(excel_file):
    wb = load_workbook(filename=excel_file, read_only=True)
    sheet_name = input('What is the name of the worksheet tab to use? ')
    sheet_ranges = wb[sheet_name]  # This defines the worksheet tab

    start_row = input('What\'s the starting row number: ')
    end_row = (sheet_ranges.max_row + 1)
    column_let = input('What column are we using for input? ')

    input_search = []
    # TODO: Rewrite this with yield
    for row in range(int(start_row), int(end_row)):
        cell_name = "{}{}".format(column_let, row)
        input_search.append(sheet_ranges[cell_name].value)
    return input_search


# If we get a hit on something, this function will take the number of hits
# and pull up all the hostnames that got a hit.
# TODO: There's an API limit we have to modify this to use
def parse_fireamp_results(query_results, search_hit):
    data_results = query_results['data']

    fireamp_hits = [
        FireampResult(**f)
        for f in data_results
    ]
    fireamp_hits.sort(key=lambda f: f.hostname)
    output_file = '{}_hosts.txt'.format(search_hit)
    with open(output_file, 'w') as result_file:
        result_file.write('The following systems got a hit on {}...\n'.format(search_hit))
        for hit in fireamp_hits:
            result_file.write('Hostname:\t{}\n'.format(hit.hostname))
    return


# If we get a hit, we redo the query using the
# total hits number as our limit value
# then we go to the function above to get a list of the hostnames
def fireamp_hostname_hits(search_term, limit):
    # If you do a requests session, you can do the authorization portion
    fireamp_req = requests.Session()
    if int(limit) < 500:
        try:
            # So this part will end up being appended to the fireamp_baseurl,
            # you can modify this line to do other searches
            # IP, Hash, etc
            fireamp_filename = "/computers/activity?q={}&limit={}".format(
                search_term, str(limit))
            # This part does the actual search
            req = fireamp_req.get(fireamp_baseurl + fireamp_filename, headers=fireamp_headers)
            # The next line keeps track of how many requests are left from FireAMP
            print('Number of requests left: {}'.format(req.headers['X-RateLimit-Remaining']))
            parse_fireamp_results(req.json(), search_term)
        except Exception as error:
            print('There was an error in the hostname query: {} Search Term: {}'.format(error, search_term))
    else:
        print('More than 500 hits, can\'t pull individual hostnames...moving on.\n')
    return


# This is the first pass parse that just queries to see if there's a hit or not
def parse_fireamp_initial_results(query_results, search_hit):
    data_results = query_results['data']
    total_results = query_results['metadata']

    if int((total_results['results']['total'])) > 0:
        print('Total hits for {}: {}'.format(search_hit,
                                             total_results['results']['total']))
        fireamp_hostname_hits(search_hit,
                              int(total_results['results']['total']))
    else:
        print('No hits on {}...moving on.'.format(search_hit))
    return


def fireamp_query_activity(search_term):
    limit = 1  # We'll add this as an option later
    # If you do a requests session, you can do the authorization portion
    fireamp_req = requests.Session()

    for item in search_term:
        try:
            # So this part will end up being appended to the fireamp_baseurl,
            # you can modify this line to do other searches
            # IP, Hash, etc
            fireamp_filename = "/computers/activity?q={}&limit={}".format(
                item, limit)
            # This part does the actual search
            req = fireamp_req.get(fireamp_baseurl + fireamp_filename, headers=fireamp_headers)
            # The next line keeps track of how many requests are left from FireAMP
            print('Number of requests left: {}'.format(req.headers['X-RateLimit-Remaining']))
            parse_fireamp_initial_results(req.json(), item)
        except Exception as error:
            print('There was an error in the initial search: {} Search Term: {}'.format(error, item))
    return


# The base URL is the starting point. Everything gets added on to that
fireamp_baseurl = "https://api.amp.cisco.com/v1/"

# The headers are recommended by the API docs
fireamp_headers = {"Accept-Encoding": "gzip, deflate",
                   "Authorization": "Basic {}".format(fireamp_base64_api),
                   "User-Agent": "FireAMP API by Python"}


def main():
    # If you want to hardcode the client ID and FireAMP API, uncomment out these lines and comment out
    # the argparse section above
    parser = argparse.ArgumentParser(description='Query the FireAMP server.')
    parser.add_argument('-f', dest='excel_file', help='Excel file for input', required=True)
    args = parser.parse_args()

    search_terms = setup_excel_file(args.excel_file)
    fireamp_results = fireamp_query_activity(search_terms)


if __name__ == '__main__':
    main()
